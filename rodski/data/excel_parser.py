"""Excel 解析器 - 读取用例、解析步骤、数据验证"""
import openpyxl
from typing import List, Dict, Any, Optional
from pathlib import Path


class ExcelParser:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None

    def load(self) -> None:
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

    def parse_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        if not self.workbook:
            self.load()
        sheet = self.workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                data.append(dict(zip(headers, row)))
        return data

    def parse(self, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        if not self.workbook:
            self.load()
        name = sheet_name or self.workbook.sheetnames[0]
        return self.parse_sheet(name)

    def get_sheet_names(self) -> List[str]:
        if not self.workbook:
            self.load()
        return self.workbook.sheetnames

    def parse_steps(self, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        rows = self.parse(sheet_name)
        steps = []
        for row in rows:
            keyword = row.get("keyword") or row.get("Keyword") or ""
            if not keyword:
                continue
            params = {}
            for k, v in row.items():
                if k and k.lower() not in ("keyword", "name", "step") and v is not None:
                    params[k.lower()] = v
            steps.append({
                "keyword": str(keyword).strip(),
                "params": params,
                "name": row.get("name") or row.get("Name") or str(keyword),
            })
        return steps

    def validate(self) -> List[str]:
        errors = []
        if not self.file_path.exists():
            errors.append(f"File not found: {self.file_path}")
            return errors
        try:
            self.load()
        except Exception as e:
            errors.append(f"Cannot load workbook: {e}")
            return errors
        for name in self.workbook.sheetnames:
            sheet = self.workbook[name]
            headers = [cell.value for cell in sheet[1]]
            if not headers or all(h is None for h in headers):
                errors.append(f"Sheet '{name}' has no headers")
        return errors

    def close(self) -> None:
        if self.workbook:
            self.workbook.close()
            self.workbook = None
