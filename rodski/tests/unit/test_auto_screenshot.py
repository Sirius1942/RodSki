"""失败自动截图功能单元测试"""
import pytest
import openpyxl
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch
from datetime import datetime

from core.ski_executor import SKIExecutor
from core.config_manager import ConfigManager
from core.result_writer import ResultWriter, HEADERS
from drivers.base_driver import BaseDriver


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def tmp_excel(tmp_path):
    """创建一个最小化的 Excel 文件（含 Case Sheet）供测试使用。"""
    path = tmp_path / "test_case.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Case"
    ws.append(["Execute", "CaseID", "Title", "PreProcess", "TestStep", "ExpectedResult", "PostProcess"])
    ws.append(["是", "TC001", "登录测试", "open||url=https://example.com", "type||locator=#input", "", ""])
    ws.append(["是", "TC002", "失败测试", "open||url=https://example.com", "click||locator=#nonexistent", "", ""])
    wb.save(path)
    return str(path)


@pytest.fixture
def tmp_model(tmp_path):
    """创建一个最小化的 XML 模型文件。"""
    path = tmp_path / "test_model.xml"
    path.write_text("""<?xml version="1.0" encoding="UTF-8"?>
<models>
  <model name="test">
    <element name="input">
      <location type="id">username</location>
    </element>
  </model>
</models>
""")
    return str(path)


@pytest.fixture
def mock_driver():
    """创建 Mock 驱动实例。"""
    driver = Mock(spec=BaseDriver)
    driver.screenshot = Mock(return_value=True)
    driver.click = Mock(return_value=True)
    driver.type = Mock(return_value=True)
    driver.navigate = Mock(return_value=True)
    driver.close = Mock(return_value=True)
    return driver


@pytest.fixture
def config_with_screenshot(tmp_path):
    """创建启用截图的配置。"""
    config = ConfigManager(str(tmp_path / "config.json"))
    config.set("auto_screenshot_on_failure", True)
    config.set("screenshot_dir", str(tmp_path / "screenshots"))
    return config


@pytest.fixture
def config_without_screenshot(tmp_path):
    """创建禁用截图的配置。"""
    config = ConfigManager(str(tmp_path / "config_no_screenshot.json"))
    config.set("auto_screenshot_on_failure", False)
    return config


# ---------------------------------------------------------------------------
# 配置测试
# ---------------------------------------------------------------------------

class TestScreenshotConfig:
    """测试截图配置开关"""
    
    def test_default_config_has_screenshot_enabled(self):
        """默认配置应启用自动截图"""
        config = ConfigManager()
        assert config.get("auto_screenshot_on_failure") == True
        assert config.get("screenshot_dir") == "screenshots"
    
    def test_config_can_disable_screenshot(self, tmp_path):
        """可以禁用自动截图"""
        config = ConfigManager(str(tmp_path / "test_config.json"))
        config.set("auto_screenshot_on_failure", False)
        assert config.get("auto_screenshot_on_failure") == False
    
    def test_config_can_set_screenshot_dir(self, tmp_path):
        """可以设置自定义截图目录"""
        config = ConfigManager(str(tmp_path / "test_config.json"))
        config.set("screenshot_dir", "custom_screenshots")
        assert config.get("screenshot_dir") == "custom_screenshots"


# ---------------------------------------------------------------------------
# SKIExecutor 截图测试
# ---------------------------------------------------------------------------

class TestSKIExecutorScreenshot:
    """测试 SKIExecutor 的自动截图功能"""
    
    def test_executor_inits_with_screenshot_config(self, tmp_excel, tmp_model, mock_driver, config_with_screenshot):
        """执行器应正确加载截图配置"""
        executor = SKIExecutor(tmp_excel, tmp_model, mock_driver, config_with_screenshot)
        assert executor.auto_screenshot == True
        assert executor.screenshot_dir.exists()
    
    def test_executor_respects_disabled_screenshot(self, tmp_excel, tmp_model, mock_driver, config_without_screenshot):
        """执行器应正确处理禁用截图配置"""
        executor = SKIExecutor(tmp_excel, tmp_model, mock_driver, config_without_screenshot)
        assert executor.auto_screenshot == False
    
    def test_screenshot_taken_on_failure(self, tmp_excel, tmp_model, mock_driver, config_with_screenshot):
        """失败时应自动截图"""
        # 设置 click 方法抛出异常
        mock_driver.click.side_effect = RuntimeError("Element not found")
        
        executor = SKIExecutor(tmp_excel, tmp_model, mock_driver, config_with_screenshot)
        
        # 模拟一个会失败的用例
        case = {
            'case_id': 'TC_FAIL',
            'title': '失败测试',
            'pre_process': {'action': ''},
            'test_step': {'action': 'click', 'model': '', 'data': ''},
            'expected_result': {'action': ''},
            'post_process': {'action': ''},
        }
        
        result = executor.execute_case(case)
        
        # 验证状态为 FAIL
        assert result['status'] == 'FAIL'
        
        # 验证截图被调用
        mock_driver.screenshot.assert_called_once()
        
        # 验证截图路径格式正确
        assert result['screenshot_path'] is not None
        assert 'TC_FAIL' in result['screenshot_path']
        assert '_failure.png' in result['screenshot_path']
    
    def test_no_screenshot_on_success(self, tmp_excel, tmp_model, mock_driver, config_with_screenshot):
        """成功时不应截图"""
        executor = SKIExecutor(tmp_excel, tmp_model, mock_driver, config_with_screenshot)
        
        # 模拟一个成功的用例
        case = {
            'case_id': 'TC_PASS',
            'title': '成功测试',
            'pre_process': {'action': ''},
            'test_step': {'action': ''},
            'expected_result': {'action': ''},
            'post_process': {'action': ''},
        }
        
        result = executor.execute_case(case)
        
        # 验证状态为 PASS
        assert result['status'] == 'PASS'
        
        # 验证截图未被调用
        mock_driver.screenshot.assert_not_called()
        
        # 验证没有截图路径
        assert result.get('screenshot_path', '') == ''
    
    def test_no_screenshot_when_disabled(self, tmp_excel, tmp_model, mock_driver, config_without_screenshot):
        """禁用截图时失败也不截图"""
        # 设置 click 方法抛出异常
        mock_driver.click.side_effect = RuntimeError("Element not found")
        
        executor = SKIExecutor(tmp_excel, tmp_model, mock_driver, config_without_screenshot)
        
        case = {
            'case_id': 'TC_FAIL_NO_SS',
            'title': '失败但不截图',
            'pre_process': {'action': ''},
            'test_step': {'action': 'click', 'model': '', 'data': ''},
            'expected_result': {'action': ''},
            'post_process': {'action': ''},
        }
        
        result = executor.execute_case(case)
        
        # 验证状态为 FAIL
        assert result['status'] == 'FAIL'
        
        # 验证截图未被调用
        mock_driver.screenshot.assert_not_called()
        
        # 验证没有截图路径
        assert result.get('screenshot_path', '') == ''
    
    def test_screenshot_filename_format(self, tmp_excel, tmp_model, mock_driver, config_with_screenshot):
        """截图文件名格式应符合规范: {case_id}_{timestamp}_failure.png"""
        mock_driver.click.side_effect = RuntimeError("Test error")
        
        executor = SKIExecutor(tmp_excel, tmp_model, mock_driver, config_with_screenshot)
        
        case = {
            'case_id': 'TC_001',
            'title': '测试文件名格式',
            'pre_process': {'action': ''},
            'test_step': {'action': 'click', 'model': '', 'data': ''},
            'expected_result': {'action': ''},
            'post_process': {'action': ''},
        }
        
        result = executor.execute_case(case)
        
        # 验证文件名格式
        path = Path(result['screenshot_path'])
        filename = path.name
        
        # 应包含 case_id
        assert 'TC_001' in filename
        # 应以 _failure.png 结尾
        assert filename.endswith('_failure.png')
        # 应包含时间戳 (YYYYMMDD_HHMMSS 格式)
        # 文件名格式: TC_001_20240319_120000_failure.png
        parts = filename.split('_')
        assert len(parts) >= 4  # TC_001, date, time, failure.png


# ---------------------------------------------------------------------------
# ResultWriter 截图路径测试
# ---------------------------------------------------------------------------

class TestResultWriterScreenshot:
    """测试 ResultWriter 对截图路径的处理"""
    
    def test_headers_include_screenshot_path(self):
        """HEADERS 应包含 ScreenshotPath 列"""
        assert "ScreenshotPath" in HEADERS
    
    def test_screenshot_path_written_to_excel(self, tmp_excel):
        """截图路径应正确写入 Excel"""
        rw = ResultWriter(tmp_excel)
        rw.write_result({
            "case_id": "TC001",
            "title": "测试截图路径写入",
            "status": "FAIL",
            "error": "Element not found",
            "screenshot_path": "screenshots/TC001_20240319_120000_failure.png"
        })
        
        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        
        # 找到 ScreenshotPath 列的位置
        headers = [sheet.cell(1, c).value for c in range(1, len(HEADERS) + 1)]
        screenshot_col = headers.index("ScreenshotPath") + 1
        
        # 验证截图路径写入正确
        assert sheet.cell(2, screenshot_col).value == "screenshots/TC001_20240319_120000_failure.png"
    
    def test_empty_screenshot_path_for_pass(self, tmp_excel):
        """通过的用例截图路径应为空"""
        rw = ResultWriter(tmp_excel)
        rw.write_result({
            "case_id": "TC001",
            "status": "PASS",
            "screenshot_path": ""
        })
        
        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        
        # 找到 ScreenshotPath 列的位置
        headers = [sheet.cell(1, c).value for c in range(1, len(HEADERS) + 1)]
        screenshot_col = headers.index("ScreenshotPath") + 1
        
        # 验证截图路径为空
        assert sheet.cell(2, screenshot_col).value in ("", None)


# ---------------------------------------------------------------------------
# 集成测试
# ---------------------------------------------------------------------------

class TestAutoScreenshotIntegration:
    """自动截图功能集成测试"""
    
    def test_full_flow_with_failure_screenshot(self, tmp_excel, tmp_model, mock_driver, config_with_screenshot):
        """测试完整流程: 失败 -> 截图 -> 写入结果"""
        mock_driver.click.side_effect = RuntimeError("Timeout waiting for element")
        
        executor = SKIExecutor(tmp_excel, tmp_model, mock_driver, config_with_screenshot)
        
        case = {
            'case_id': 'TC_INTEGRATION',
            'title': '集成测试',
            'pre_process': {'action': ''},
            'test_step': {'action': 'click', 'model': '', 'data': ''},
            'expected_result': {'action': ''},
            'post_process': {'action': ''},
        }
        
        result = executor.execute_case(case)
        
        # 验证结果完整性
        assert result['case_id'] == 'TC_INTEGRATION'
        assert result['status'] == 'FAIL'
        assert 'Timeout' in result['error']
        assert result['screenshot_path'] != ''
        
        # 验证截图文件存在（Mock 情况下路径生成正确）
        path = Path(result['screenshot_path'])
        assert path.name.startswith('TC_INTEGRATION')
        assert path.name.endswith('_failure.png')
    
    def test_screenshot_failure_does_not_break_execution(self, tmp_excel, tmp_model, mock_driver, config_with_screenshot):
        """截图失败不应中断测试执行"""
        # 设置截图方法抛出异常
        mock_driver.screenshot.side_effect = Exception("Screenshot failed")
        mock_driver.click.side_effect = RuntimeError("Test error")
        
        executor = SKIExecutor(tmp_excel, tmp_model, mock_driver, config_with_screenshot)
        
        case = {
            'case_id': 'TC_SCREENSHOT_FAIL',
            'title': '截图失败测试',
            'pre_process': {'action': ''},
            'test_step': {'action': 'click', 'model': '', 'data': ''},
            'expected_result': {'action': ''},
            'post_process': {'action': ''},
        }
        
        # 不应抛出异常
        result = executor.execute_case(case)
        
        # 验证结果仍然正确
        assert result['status'] == 'FAIL'
        # 截图路径应为空（因为截图失败）
        assert result['screenshot_path'] == ''