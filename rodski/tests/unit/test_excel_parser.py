"""ExcelParser 单元测试"""
import pytest
import openpyxl
from pathlib import Path
from data.excel_parser import ExcelParser


@pytest.fixture
def sample_xlsx(tmp_path):
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestCases"
    ws.append(["keyword", "name", "locator", "text"])
    ws.append(["click", "Click button", "#btn", None])
    ws.append(["type", "Type text", "#input", "hello"])
    ws.append(["wait", "Wait 2s", None, None])
    wb.save(path)
    return str(path)


@pytest.fixture
def empty_xlsx(tmp_path):
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty"
    ws.append(["keyword", "name"])
    wb.save(path)
    return str(path)


class TestExcelParser:
    def test_load(self, sample_xlsx):
        parser = ExcelParser(sample_xlsx)
        parser.load()
        assert parser.workbook is not None
        parser.close()

    def test_load_nonexistent(self):
        parser = ExcelParser("/nonexistent/file.xlsx")
        with pytest.raises(FileNotFoundError):
            parser.load()

    def test_parse_sheet(self, sample_xlsx):
        parser = ExcelParser(sample_xlsx)
        data = parser.parse_sheet("TestCases")
        assert len(data) == 3
        assert data[0]["keyword"] == "click"
        assert data[1]["text"] == "hello"
        parser.close()

    def test_parse_default_sheet(self, sample_xlsx):
        parser = ExcelParser(sample_xlsx)
        data = parser.parse()
        assert len(data) == 3
        parser.close()

    def test_get_sheet_names(self, sample_xlsx):
        parser = ExcelParser(sample_xlsx)
        names = parser.get_sheet_names()
        assert "TestCases" in names
        parser.close()

    def test_parse_steps(self, sample_xlsx):
        parser = ExcelParser(sample_xlsx)
        steps = parser.parse_steps()
        assert len(steps) == 3
        assert steps[0]["keyword"] == "click"
        assert steps[0]["params"]["locator"] == "#btn"
        assert steps[1]["params"]["text"] == "hello"
        parser.close()

    def test_parse_empty_sheet(self, empty_xlsx):
        parser = ExcelParser(empty_xlsx)
        data = parser.parse()
        assert data == []
        parser.close()

    def test_validate_valid(self, sample_xlsx):
        parser = ExcelParser(sample_xlsx)
        errors = parser.validate()
        assert errors == []
        parser.close()

    def test_validate_missing_file(self):
        parser = ExcelParser("/nonexistent/file.xlsx")
        errors = parser.validate()
        assert len(errors) == 1
        assert "not found" in errors[0].lower() or "File not found" in errors[0]

    def test_close(self, sample_xlsx):
        parser = ExcelParser(sample_xlsx)
        parser.load()
        parser.close()
        assert parser.workbook is None

    def test_parse_skips_empty_rows(self, tmp_path):
        path = tmp_path / "gaps.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["keyword", "name"])
        ws.append(["click", "Step 1"])
        ws.append([None, None])
        ws.append(["type", "Step 2"])
        wb.save(path)
        parser = ExcelParser(str(path))
        data = parser.parse()
        assert len(data) == 2  # empty row [None, None] is skipped
        parser.close()

    def test_parse_steps_skips_non_keyword(self, tmp_path):
        path = tmp_path / "nokey.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["keyword", "name", "locator"])
        ws.append([None, "No keyword", "#btn"])
        ws.append(["click", "Has keyword", "#btn"])
        wb.save(path)
        parser = ExcelParser(str(path))
        steps = parser.parse_steps()
        assert len(steps) == 1
        assert steps[0]["keyword"] == "click"
        parser.close()
