"""ResultWriter 单元测试"""
import pytest
import openpyxl
from pathlib import Path
from unittest.mock import patch
from core.result_writer import ResultWriter, HEADERS


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def tmp_excel(tmp_path):
    """创建一个最小化的 Excel 文件（含 Case Sheet）供测试使用。"""
    path = tmp_path / "test_case.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Case"
    ws.append(["Execute", "CaseID", "Title"])
    ws.append(["是", "TC001", "登录测试"])
    ws.append(["是", "TC002", "搜索测试"])
    wb.save(path)
    return str(path)


def _load_result_sheet(path: str):
    wb = openpyxl.load_workbook(path)
    assert "TestResult" in wb.sheetnames
    sheet = wb["TestResult"]
    headers = [sheet.cell(1, c).value for c in range(1, len(HEADERS) + 1)]
    rows = []
    for r in range(2, sheet.max_row + 1):
        row = {headers[c]: sheet.cell(r, c + 1).value for c in range(len(HEADERS))}
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# 基础功能
# ---------------------------------------------------------------------------

class TestResultWriterInit:
    def test_raises_if_file_missing(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            ResultWriter(str(tmp_path / "nonexistent.xlsx"))

    def test_init_ok(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        assert rw.case_file == Path(tmp_excel)


class TestSheetCreation:
    def test_creates_testresult_sheet(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "status": "PASS"})

        wb = openpyxl.load_workbook(tmp_excel)
        assert "TestResult" in wb.sheetnames

    def test_headers_correct(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "status": "PASS"})

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        actual = [sheet.cell(1, c).value for c in range(1, len(HEADERS) + 1)]
        assert actual == HEADERS

    def test_does_not_duplicate_sheet(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "status": "PASS"})
        rw.write_result({"case_id": "TC002", "status": "FAIL"})

        wb = openpyxl.load_workbook(tmp_excel)
        assert wb.sheetnames.count("TestResult") == 1


# ---------------------------------------------------------------------------
# 写入内容
# ---------------------------------------------------------------------------

class TestWriteResult:
    def test_pass_result_written(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "title": "登录测试", "status": "PASS", "execution_time": 1.23})

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        assert sheet.cell(2, 1).value == "TC001"
        assert sheet.cell(2, 3).value == "PASS"
        assert sheet.cell(2, 4).value == 1.23

    def test_fail_result_with_error(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({
            "case_id": "TC002",
            "title": "搜索测试",
            "status": "FAIL",
            "execution_time": 0.5,
            "error": "Element not found",
        })

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        assert sheet.cell(2, 3).value == "FAIL"
        assert sheet.cell(2, 9).value == "Element not found"

    def test_updated_at_populated(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "status": "PASS"})

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        assert sheet.cell(2, 11).value is not None


# ---------------------------------------------------------------------------
# 批量写入
# ---------------------------------------------------------------------------

class TestWriteResults:
    def test_batch_write(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        results = [
            {"case_id": "TC001", "title": "登录测试", "status": "PASS", "execution_time": 1.0},
            {"case_id": "TC002", "title": "搜索测试", "status": "FAIL", "execution_time": 0.8, "error": "Timeout"},
        ]
        rw.write_results(results)

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        assert sheet.max_row == 3  # header + 2 data rows

    def test_empty_list_no_error(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_results([])  # should not raise


# ---------------------------------------------------------------------------
# 更新（upsert）行为
# ---------------------------------------------------------------------------

class TestUpsert:
    def test_existing_case_updated(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "status": "FAIL", "error": "first run"})
        rw.write_result({"case_id": "TC001", "status": "PASS"})

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        # 只应有一行数据
        assert sheet.max_row == 2
        assert sheet.cell(2, 3).value == "PASS"
        assert sheet.cell(2, 9).value in ("", None)  # error 清空

    def test_new_case_appended(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "status": "PASS"})
        rw.write_result({"case_id": "TC003", "status": "PASS"})

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        assert sheet.max_row == 3


# ---------------------------------------------------------------------------
# 颜色填充
# ---------------------------------------------------------------------------

class TestCellFill:
    def test_pass_green_fill(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "status": "PASS"})

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        fill = sheet.cell(2, 3).fill
        assert fill.fgColor.rgb.endswith("C6EFCE")

    def test_fail_red_fill(self, tmp_excel):
        rw = ResultWriter(tmp_excel)
        rw.write_result({"case_id": "TC001", "status": "FAIL"})

        wb = openpyxl.load_workbook(tmp_excel)
        sheet = wb["TestResult"]
        fill = sheet.cell(2, 3).fill
        assert fill.fgColor.rgb.endswith("FFC7CE")
