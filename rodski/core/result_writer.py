"""结果回填模块 - 将测试执行结果写入 Excel 的 TestResult Sheet"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from collections import Counter

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("rodski")

# 列定义 - 扩展版本
HEADERS = [
    "CaseID", "Title", "Status", "ExecutionTime", 
    "Retries", "StartTime", "EndTime", 
    "ErrorType", "ErrorMessage", "ScreenshotPath", "UpdatedAt"
]

# 颜色定义
_COLORS = {
    "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "SKIP": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "ERROR": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

# 字体样式
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExecutionSummary:
    """执行结果统计"""
    
    def __init__(self):
        self.total: int = 0
        self.passed: int = 0
        self.failed: int = 0
        self.skipped: int = 0
        self.errors: int = 0
        self.total_time: float = 0.0
        self.error_types: Counter = Counter()
        self.start_time: Optional[datetime] = None
        self.end_time: Optional[datetime] = None
    
    def add_result(self, result: Dict[str, Any]) -> None:
        """添加单个结果到统计"""
        self.total += 1
        status = result.get("status", "FAIL").upper()
        
        if status == "PASS":
            self.passed += 1
        elif status == "SKIP":
            self.skipped += 1
        elif status == "ERROR":
            self.errors += 1
        else:
            self.failed += 1
        
        exec_time = result.get("execution_time", 0)
        if isinstance(exec_time, (int, float)):
            self.total_time += exec_time
        
        error_type = result.get("error_type", "")
        if error_type:
            self.error_types[error_type] += 1
    
    @property
    def pass_rate(self) -> float:
        """计算通过率"""
        if self.total == 0:
            return 0.0
        return (self.passed / self.total) * 100
    
    @property
    def average_time(self) -> float:
        """计算平均执行时间"""
        if self.total == 0:
            return 0.0
        return self.total_time / self.total
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "total": self.total,
            "passed": self.passed,
            "failed": self.failed,
            "skipped": self.skipped,
            "errors": self.errors,
            "pass_rate": f"{self.pass_rate:.1f}%",
            "total_time": f"{self.total_time:.2f}s",
            "average_time": f"{self.average_time:.2f}s",
            "error_types": dict(self.error_types),
            "start_time": self.start_time.strftime("%Y-%m-%d %H:%M:%S") if self.start_time else None,
            "end_time": self.end_time.strftime("%Y-%m-%d %H:%M:%S") if self.end_time else None,
        }


class ResultWriter:
    """将测试结果回填到原 Excel 文件的 TestResult Sheet。"""

    SHEET_NAME = "TestResult"
    SUMMARY_SHEET_NAME = "TestSummary"

    def __init__(self, case_file: str):
        self.case_file = Path(case_file)
        if not self.case_file.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {self.case_file}")
        self._summary = ExecutionSummary()

    # ------------------------------------------------------------------
    # 公开接口
    # ------------------------------------------------------------------

    def write_result(self, result: Dict[str, Any]) -> None:
        """写入单条用例结果。"""
        self.write_results([result])

    def write_results(self, results: List[Dict[str, Any]]) -> None:
        """批量写入用例结果。"""
        if not results:
            return
        
        # 更新统计
        self._summary.start_time = datetime.now()
        for result in results:
            self._summary.add_result(result)
        self._summary.end_time = datetime.now()
        
        # 写入结果
        wb = openpyxl.load_workbook(self.case_file)
        sheet = self._get_or_create_sheet(wb)
        index = self._build_index(sheet)

        for result in results:
            self._upsert_row(sheet, index, result)
        
        # 写入统计摘要
        self._write_summary_sheet(wb)

        wb.save(self.case_file)
        
        # 输出日志
        summary = self._summary.to_dict()
        logger.info(
            f"结果已回填到 {self.case_file.name}: "
            f"总计 {summary['total']} 条, "
            f"通过 {summary['passed']} 条, "
            f"失败 {summary['failed']} 条, "
            f"通过率 {summary['pass_rate']}"
        )

    def get_summary(self) -> ExecutionSummary:
        """获取执行统计"""
        return self._summary

    # ------------------------------------------------------------------
    # 内部方法
    # ------------------------------------------------------------------

    def _get_or_create_sheet(self, wb: openpyxl.Workbook):
        if self.SHEET_NAME in wb.sheetnames:
            return wb[self.SHEET_NAME]

        sheet = wb.create_sheet(self.SHEET_NAME)
        self._write_headers(sheet)
        return sheet

    def _write_headers(self, sheet) -> None:
        """写入表头"""
        for col, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER
            # 设置列宽
            col_letter = get_column_letter(col)
            if header == "CaseID":
                sheet.column_dimensions[col_letter].width = 10
            elif header == "Title":
                sheet.column_dimensions[col_letter].width = 30
            elif header == "ErrorMessage":
                sheet.column_dimensions[col_letter].width = 50
            elif header == "ScreenshotPath":
                sheet.column_dimensions[col_letter].width = 40
            else:
                sheet.column_dimensions[col_letter].width = 15

    def _build_index(self, sheet) -> Dict[str, int]:
        """构建 CaseID -> 行号 的索引（跳过表头）。"""
        index = {}
        for row_idx in range(2, sheet.max_row + 1):
            case_id = sheet.cell(row=row_idx, column=1).value
            if case_id:
                index[str(case_id)] = row_idx
        return index

    def _upsert_row(self, sheet, index: Dict[str, int], result: Dict[str, Any]) -> None:
        case_id = str(result.get("case_id", ""))
        title = str(result.get("title", ""))
        status = str(result.get("status", "FAIL")).upper()
        exec_time = result.get("execution_time", "")
        retries = result.get("retries", 0)
        start_time = result.get("start_time", "")
        end_time = result.get("end_time", "")
        error_type = result.get("error_type", "")
        error_msg = str(result.get("error", ""))
        screenshot_path = str(result.get("screenshot_path", ""))
        updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if case_id in index:
            row_idx = index[case_id]
        else:
            row_idx = sheet.max_row + 1
            index[case_id] = row_idx

        values = [
            case_id, title, status, exec_time,
            retries, start_time, end_time,
            error_type, error_msg, screenshot_path, updated
        ]
        fill = _COLORS.get(status, _COLORS["FAIL"])

        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")
    
    def _write_summary_sheet(self, wb: openpyxl.Workbook) -> None:
        """写入统计摘要 Sheet"""
        if self.SUMMARY_SHEET_NAME in wb.sheetnames:
            sheet = wb[self.SUMMARY_SHEET_NAME]
            # 清空现有内容
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            sheet = wb.create_sheet(self.SUMMARY_SHEET_NAME)
        
        summary = self._summary.to_dict()
        
        # 写入统计数据
        rows = [
            ["指标", "值"],
            ["总用例数", summary["total"]],
            ["通过数", summary["passed"]],
            ["失败数", summary["failed"]],
            ["跳过数", summary["skipped"]],
            ["错误数", summary["errors"]],
            ["通过率", summary["pass_rate"]],
            ["总执行时间", summary["total_time"]],
            ["平均执行时间", summary["average_time"]],
            ["开始时间", summary["start_time"]],
            ["结束时间", summary["end_time"]],
        ]
        
        # 添加错误类型统计
        if summary["error_types"]:
            rows.append(["", ""])
            rows.append(["错误类型统计", ""])
            for error_type, count in summary["error_types"].items():
                rows.append([error_type, count])
        
        for row_idx, row_data in enumerate(rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or (row_idx == 12 and col_idx == 1):
                    cell.font = Font(bold=True)
                cell.border = _BORDER
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 25
