"""Case Sheet 解析器 - 解析三段式用例结构"""
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook


class CaseParser:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path, data_only=True)

    def parse_cases(self) -> List[Dict[str, Any]]:
        """解析 Case Sheet"""
        if 'Case' not in self.wb.sheetnames:
            return []

        sheet = self.wb['Case']
        cases = []

        for row_idx in range(3, sheet.max_row + 1):
            row = sheet[row_idx]
            execute_control = str(row[0].value or '').strip()

            if execute_control != '是':
                continue

            case = {
                'case_id': str(row[1].value or ''),
                'title': str(row[2].value or ''),
                'description': str(row[3].value or ''),
                'component_type': str(row[4].value or ''),
                'pre_process': {
                    'action': str(row[5].value or ''),
                    'model': str(row[6].value or ''),
                    'data': str(row[7].value or '')
                },
                'test_step': {
                    'action': str(row[8].value or ''),
                    'model': str(row[9].value or ''),
                    'data': str(row[10].value or '')
                },
                'expected_result': {
                    'action': str(row[11].value or ''),
                    'model': str(row[12].value or ''),
                    'data': str(row[13].value or '')
                },
                'post_process': {
                    'action': str(row[14].value or ''),
                    'model': str(row[15].value or ''),
                    'data': str(row[16].value or '')
                }
            }
            cases.append(case)

        return cases

    def close(self):
        self.wb.close()
