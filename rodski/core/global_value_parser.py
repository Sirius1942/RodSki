"""全局变量解析器 - 解析 GlobalValue Sheet"""
from typing import Dict, Any
from openpyxl import load_workbook


class GlobalValueParser:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path, data_only=True)

    def parse(self) -> Dict[str, Dict[str, str]]:
        """解析全局变量，返回 {分组: {变量名: 值}}"""
        if 'GlobalValue' not in self.wb.sheetnames:
            return {}

        sheet = self.wb['GlobalValue']
        global_vars = {}

        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            group = str(row[0].value or '').strip()
            var_name = str(row[1].value or '').strip()
            value = str(row[2].value or '').strip()

            if not group or not var_name:
                continue

            if group not in global_vars:
                global_vars[group] = {}

            global_vars[group][var_name] = value

        return global_vars

    def close(self):
        self.wb.close()
