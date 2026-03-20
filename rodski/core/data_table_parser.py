"""数据表解析器 - 解析 DataID/Remark/动态列"""
from typing import Dict, Any
from openpyxl import load_workbook


class DataTableParser:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path, data_only=True)
        self.tables = {}

    def parse_all_tables(self) -> Dict[str, Dict[str, Dict[str, Any]]]:
        """解析所有数据表（除了 Main/Case/GlobalValue/TestResult/Logic）"""
        skip_sheets = {'Main', 'Case', 'GlobalValue', 'TestResult'}

        for sheet_name in self.wb.sheetnames:
            if sheet_name not in skip_sheets:
                self.tables[sheet_name] = self.parse_table(sheet_name)

        return self.tables

    def parse_table(self, sheet_name: str) -> Dict[str, Dict[str, Any]]:
        """解析单个数据表"""
        if sheet_name not in self.wb.sheetnames:
            return {}

        sheet = self.wb[sheet_name]
        headers = [str(cell.value or '') for cell in sheet[1]]

        if not headers or headers[0] != 'DataID':
            return {}

        field_names = headers[2:]
        table_data = {}

        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data_id = str(row[0].value or '').strip()

            if not data_id:
                continue

            row_data = {}
            for i, field_name in enumerate(field_names):
                cell_value = row[i + 2].value
                if cell_value and str(cell_value).strip():
                    row_data[field_name] = str(cell_value).strip()

            table_data[data_id] = row_data

        return table_data

    def get_data(self, table_name: str, data_id: str) -> Dict[str, Any]:
        """获取指定数据行"""
        return self.tables.get(table_name, {}).get(data_id, {})

    def close(self):
        self.wb.close()
